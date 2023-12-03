import openpyxl


# row count
def getrowCount(file, sheetname):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return (sheet.max_row)

# read data
def readData(file, sheetname, rowno, colno):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.cell(row=rowno, column=colno).value

# write data
def writeDate(file, sheetname, rowno, colno, data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    sheet.cell(row=rowno, column=colno).value=data
    book.save(file)
